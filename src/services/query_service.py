from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from src.llm.sql_agent import SQLAgent
from src.llm.summarizer import ResultSummarizer
from src.db.schema_inspect import SchemaInspector
from src.core.logging import get_logger
from src.core.config import settings
from pathlib import Path
from typing import Dict, Optional, AsyncGenerator
from src.utils.json_safe import make_json_safe

import time
import uuid
import json
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = get_logger(__name__)


class QueryService:
    def __init__(self, session: AsyncSession, session_id: str = "default", user_id: Optional[int] = None):
        self.session = session
        self.session_id = session_id
        self.user_id = user_id
        self.sql_agent = SQLAgent(session_id=session_id)
        self.summarizer = ResultSummarizer(session_id=f"{session_id}_summ")
        self.schema_inspector = SchemaInspector(session)
        self.schema_context = SchemaInspector._cached_schema_text or ""

    # ===============================================================
    #  Main Stream Process (Intent Detection + Execution)
    # ===============================================================
    async def stream_process(self, user_query: str) -> AsyncGenerator[Dict, None]:
        start_time = time.time()
        sql = ""

        try:
            await self._save_chat_message("user", user_query)
            yield {"event": "status", "message": "🧠 Understanding your intent..."}
            intent = await self._detect_intent(user_query)
            logger.info(f"🧠 Detected intent: {intent}")

            # =======================================================
            # 1️⃣ CHAT
            # =======================================================
            if intent == "CHAT":
                yield {"event": "status", "message": "💬 Generating conversational reply..."}
                try:
                    response = await self.summarizer.generate_chat_reply(user_query)
                except Exception as e:
                    logger.error(f"Chat reply error: {e}")
                    response = None

                response = response or "Hey there! How can I assist you with your data today?"
                await self._save_chat_message("assistant", response)
                yield {"event": "chat_reply", "message": response}
                await self._log_query(user_query, None, "valid", "success", None, 0, int((time.time() - start_time) * 1000))
                return

            # =======================================================
            # 2️⃣ DATA_SUMMARY — Full Dataset Analysis (new)
            # =======================================================
            # =======================================================
# 2️⃣ DATA_SUMMARY — Full Dataset Analysis (new)
# =======================================================
            if intent == "DATA_SUMMARY":
                yield {"event": "status", "message": "📊 Analyzing your database structure and contents..."}

                try:
                    # Get database schema + sample data
                    schema_text = await self.schema_inspector.get_schema_description()
                    total_rows_query = text(f"SELECT COUNT(*) FROM {settings.SCHEMA}.{settings.TABLE_NAME}")
                    total_rows = (await self.session.execute(total_rows_query)).scalar() or 0

                    sample_query = text(f"SELECT * FROM {settings.SCHEMA}.{settings.TABLE_NAME} LIMIT 200")
                    result = await self.session.execute(sample_query)
                    sample_data = [dict(r) for r in result.mappings().all()]

                    if not sample_data:
                        yield {"event": "error", "message": "⚠️ No records found in the table."}
                        return

                    # Build DataFrame and generate profiling
                    df = pd.DataFrame(sample_data)
                    profile = []
                    for col in df.columns:
                        col_data = df[col]
                        profile.append({
                            "column": col,
                            "dtype": str(col_data.dtype),
                            "unique_values": int(col_data.nunique(dropna=True)),
                            "missing_values": int(col_data.isna().sum()),
                            "sample_values": list(col_data.dropna().unique()[:5]),
                            "min": float(col_data.min()) if pd.api.types.is_numeric_dtype(col_data) else None,
                            "max": float(col_data.max()) if pd.api.types.is_numeric_dtype(col_data) else None,
                            "mean": float(col_data.mean()) if pd.api.types.is_numeric_dtype(col_data) else None,
                        })

                    # Create structured summary
                    profile_summary = {
                        "total_rows": int(total_rows),
                        "total_columns": len(df.columns),
                        "column_profiles": profile,
                    }

                    # Generate AI summary
                    prompt = f"""
            You are a data analyst. Analyze the following database and summarize it.
            Describe:
            - What this data represents
            - Key columns and their meanings
            - Total records and variations (unique values)
            - Numeric column ranges and averages
            - Any insights or patterns you notice

            Schema:
            {schema_text}

            Profile Statistics:
            {json.dumps(profile_summary, indent=2, default=str)}

            Write a 5–6 line descriptive summary.
            """
                    summary_text = await self.summarizer.llm_provider.generate_response(prompt)
                    summary_text = summary_text.strip() or "This database contains structured tabular data for analysis."

                    # Convert to JSON-safe format before yielding
                    safe_profile = make_json_safe(profile_summary)

                    await self._save_chat_message("assistant", summary_text)
                    yield {
                        "event": "data_profile",
                        "message": "Here’s a detailed overview of your database:",
                        "data": {
                            "summary_text": summary_text,
                            "schema": schema_text,
                            "profile": safe_profile,
                        },
                    }
                    return

                except Exception as e:
                    logger.error(f"DATA_SUMMARY failed: {e}")
                    yield {"event": "error", "message": f"Failed to analyze database: {e}"}
                    await self.session.rollback()
                    return


            # =======================================================
            # 3️⃣ OTHER Intent
            # =======================================================
            if intent == "OTHER":
                reply = "🤔 I couldn’t understand that clearly. Could you rephrase or ask about your data?"
                await self._save_chat_message("assistant", reply)
                yield {"event": "chat_reply", "message": reply}
                return

            # =======================================================
            # 4️⃣ SQL_QUERY
            # =======================================================
            yield {"event": "status", "message": "🧩 Generating SQL query from your request..."}
            schema_info = self.schema_context or await self.schema_inspector.get_schema_description()

            async for piece in self.sql_agent.stream_generate_sql(user_query):
                if isinstance(piece, str):
                    yield {"event": "sql_token", "chunk": piece}
                    sql += piece
                elif isinstance(piece, dict) and piece.get("__final_sql__"):
                    sql = piece["__final_sql__"]
                    yield {"event": "sql", "sql": sql}

            if not sql.strip():
                yield {"event": "error", "message": "❌ No valid SQL generated."}
                return

            valid, fixed_sql, error = self.sql_agent.validate_and_fix_sql(sql)
            if not valid:
                await self._log_query(user_query, sql, "invalid", "failed", error)
                yield {"event": "error", "message": error or "Invalid SQL syntax."}
                return

            yield {"event": "status", "message": "🚀 Executing SQL query on the database..."}

            result = await self.session.execute(text(fixed_sql))
            rows = result.fetchall()
            columns = list(result.keys())
            if not rows:
                yield {"event": "error", "message": "⚠️ No records found for this query."}
                return

            rows_list = [list(r) for r in rows]
            df = pd.DataFrame(rows_list, columns=columns).dropna(how="all")
            df.columns = [str(c).strip() for c in df.columns]
            total_rows = len(df)

            file_name = await self._export_to_excel(df)
            preview_html = self._generate_preview_html(df, total_rows, file_name)
            yield {"event": "rows_preview", "data": {
                "columns": columns,
                "preview_html": preview_html,
                "row_count": total_rows,
                "file_name": file_name
            }}

            yield {"event": "status", "message": "🧠 Generating AI summary..."}

            rows_tuples = [tuple(r) for r in rows]
            summary_text = await self.summarizer.summarize(user_query, fixed_sql, columns, rows_tuples, total_rows)
            summary_text = summary_text or f"Query executed successfully with {total_rows} rows."

            await self._save_chat_message("assistant", summary_text)
            await self._update_session_summary(summary_text)

            exec_time_ms = int((time.time() - start_time) * 1000)
            await self._log_query(user_query, fixed_sql, "valid", "success", None, total_rows, exec_time_ms)
            yield {"event": "complete", "data": {
                "summary_html": f"<div><strong>AI Summary:</strong><p>{summary_text}</p></div>",
                "execution_time_ms": exec_time_ms,
                "row_count": total_rows,
                "sql": fixed_sql,
                "download_url": f"/downloads/{file_name}"
            }}

        except Exception as e:
            logger.error(f"Stream process error: {e}")
            await self.session.rollback()
            yield {"event": "error", "message": str(e)}

    # ===============================================================
    #  Intent Detection
    # ===============================================================
    async def _detect_intent(self, query: str) -> str:
        q_lower = query.lower().strip()

        smalltalk = ["hi", "hello", "hey", "thanks", "who are you", "how are you"]
        data_summary = [
            "summary", "describe", "explain", "information", "details",
            "about database", "about data", "dataset", "what is this data",
            "variations", "types", "how many types", "total types"
        ]
        sql_triggers = [
            "show", "list", "find", "fetch", "get", "select",
            "count", "total", "average", "sum", "highest", "lowest",
            "details", "data", "customer", "product", "sales", "item", "price",
            "report", "range", "tell me"
        ]

        if any(k in q_lower for k in smalltalk):
            return "CHAT"
        if any(k in q_lower for k in data_summary):
            return "DATA_SUMMARY"
        if any(k in q_lower for k in sql_triggers):
            return "SQL_QUERY"
        return "OTHER"

    # ===============================================================
    #  Excel Export + HTML Preview
    # ===============================================================
    async def _export_to_excel(self, df: pd.DataFrame) -> str:
        exports_dir = Path("exports")
        exports_dir.mkdir(exist_ok=True)
        file_name = f"report_{self.session_id}_{int(time.time())}.xlsx"
        file_path = exports_dir / file_name

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Results")
            sheet = writer.book["Results"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border = Border(
                left=Side(border_style="thin", color="D3D3D3"),
                right=Side(border_style="thin", color="D3D3D3"),
                top=Side(border_style="thin", color="D3D3D3"),
                bottom=Side(border_style="thin", color="D3D3D3"),
            )

            for i, col_name in enumerate(df.columns, 1):
                cell = sheet.cell(row=1, column=i)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                sheet.column_dimensions[get_column_letter(i)].width = max(15, len(col_name) + 2)

        return file_name

    def _generate_preview_html(self, df: pd.DataFrame, total_rows: int, file_name: str) -> str:
        preview_df = df.head(10)
        html_table = preview_df.to_html(classes="table table-striped table-bordered", index=False, border=0)
        return f"""
        <div style='margin-top:10px;background:#f9f9f9;border-radius:8px;padding:10px;'>
            <h4>Preview (Top 10 of {total_rows} rows)</h4>
            {html_table}
            <div style='margin-top:10px;text-align:center;'>
                <button onclick="window.open('/downloads/{file_name}','_blank')"
                    style='background:#0078D4;color:white;padding:8px 15px;border:none;border-radius:6px;cursor:pointer;'>
                    Download Excel (.xlsx)
                </button>
            </div>
        </div>
        """

    # ===============================================================
    #  Logging + History
    # ===============================================================
    async def _save_chat_message(self, role: str, content: str):
        message_id = str(uuid.uuid4())
        await self.session.execute(
            text("""
                INSERT INTO public.chat_history (session_id, message_id, role, content, meta_data)
                VALUES (:sid, :mid, :r, :c, :meta)
            """),
            {"sid": self.session_id, "mid": message_id, "r": role, "c": content, "meta": json.dumps({"source": "ws"})},
        )
        await self.session.commit()

    async def _update_session_summary(self, summary: str):
        await self.session.execute(
            text("""
                INSERT INTO public.session_summaries (session_id, user_id, summary)
                VALUES (:sid, NULL, :summary)
                ON CONFLICT (session_id, user_id)
                DO UPDATE SET summary = :summary, updated_at = CURRENT_TIMESTAMP
            """),
            {"sid": self.session_id, "summary": summary},
        )
        await self.session.commit()

    async def _log_query(self, user_query, generated_sql=None, validation_status=None,
                         execution_status=None, error_message=None, row_count=None, execution_time_ms=None):
        await self.session.execute(
            text("""
                INSERT INTO public.query_logs
                (session_id, user_query, generated_sql, validation_status, execution_status, error_message, row_count, execution_time_ms)
                VALUES (:sid, :uq, :gsql, :v, :e, :err, :rows, :time)
            """),
            {"sid": self.session_id, "uq": user_query, "gsql": generated_sql, "v": validation_status,
             "e": execution_status, "err": error_message, "rows": row_count, "time": execution_time_ms},
        )
        await self.session.commit()

    async def get_chat_history(self, limit: int = 50):
        """
        Fetch chat messages for the session (default: last 50).
        """
        try:
            result = await self.session.execute(
                text("""
                    SELECT role, content
                    FROM public.chat_history
                    WHERE session_id = :sid
                    ORDER BY created_at DESC
                    LIMIT :limit
                """),
                {"sid": self.session_id, "limit": limit}
            )
            rows = result.mappings().all()
            history = [{"role": r["role"], "content": r["content"]} for r in rows]
            return list(reversed(history))
        except Exception as e:
            logger.error(f"Error fetching chat history: {e}")
            return []

    async def clear_history(self) -> bool:
        """
        Delete all chat history for the current session.
        """
        try:
            await self.session.execute(
                text("DELETE FROM public.chat_history WHERE session_id = :sid"),
                {"sid": self.session_id},
            )
            await self.session.commit()
            logger.info(f"🧹 Cleared chat history for session: {self.session_id}")
            return True
        except Exception as e:
            logger.error(f"Error clearing chat history: {e}")
            return False



# from sqlalchemy import text
# from sqlalchemy.ext.asyncio import AsyncSession
# from src.llm.sql_agent import SQLAgent
# from src.llm.summarizer import ResultSummarizer
# from src.db.schema_inspect import SchemaInspector
# from src.core.logging import get_logger
# from pathlib import Path
# from typing import Dict, Optional, AsyncGenerator
# import time
# import uuid
# import json
# import pandas as pd
# from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
# from openpyxl.utils import get_column_letter

# logger = get_logger(__name__)

# class QueryService:
#     def __init__(self, session: AsyncSession, session_id: str = "default", user_id: Optional[int] = None):
#         self.session = session
#         self.session_id = session_id
#         self.user_id = user_id
#         self.sql_agent = SQLAgent(session_id=session_id)
#         self.summarizer = ResultSummarizer(session_id=f"{session_id}_summ")
#         self.schema_inspector = SchemaInspector(session)

#     # ============================================================
#     # 🎯 Main Query Processing (Unlimited Rows + Styled Excel)
#     # ============================================================
#     async def stream_process(self, user_query: str) -> AsyncGenerator[Dict, None]:
#         """
#         Full intelligent query processor:
#         - Understands intent
#         - Generates, validates, and executes SQL
#         - Exports all rows (no limit)
#         - Creates styled Excel (.xlsx)
#         - Shows top 10 preview + Download button
#         """
#         start_time = time.time()
#         sql = ""

#         try:
#             # Step 1: Log user query
#             await self._save_chat_message("user", user_query)
#             yield {"event": "status", "message": "🧠 Understanding your question and generating SQL..."}

#             # Step 2: Load DB schema for LLM
#             schema_info = await self.schema_inspector.get_schema_description()

#             # Step 3: Generate SQL using LLM (streaming)
#             async for piece in self.sql_agent.stream_generate_sql(user_query, schema_info):
#                 if isinstance(piece, str):
#                     yield {"event": "sql_token", "chunk": piece}
#                     sql += piece
#                 elif isinstance(piece, dict) and piece.get("__final_sql__"):
#                     sql = piece["__final_sql__"]
#                     yield {"event": "sql", "sql": sql}

#             if not sql.strip():
#                 yield {"event": "error", "message": "❌ Failed to generate SQL from your query."}
#                 return

#             # Step 4: Validate SQL
#             valid, fixed_sql, error = self.sql_agent.validate_and_fix_sql(sql)
#             if not valid:
#                 await self._log_query(user_query, sql, "invalid", "failed", error)
#                 yield {"event": "error", "message": error}
#                 return

#             yield {"event": "status", "message": "🚀 Running your SQL query (fetching all rows)..."}

#             # Step 5: Execute SQL (⚡ no limit)
#             result = await self.session.execute(text(fixed_sql))
#             rows = result.fetchall()
#             columns = list(result.keys())

#             if not rows:
#                 yield {"event": "error", "message": "No data found for your query."}
#                 return

#             # Step 6: Convert results to DataFrame
#             rows_list = [list(r) for r in rows]
#             df = pd.DataFrame(rows_list, columns=columns)
#             df = df.dropna(how="all")
#             df.columns = [str(c).strip() for c in df.columns]
#             total_rows = len(df)

#             # Step 7: Export to styled Excel (manual download)
#             exports_dir = Path("exports")
#             exports_dir.mkdir(exist_ok=True)
#             file_name = f"report_{self.session_id}_{int(time.time())}.xlsx"
#             file_path = exports_dir / file_name

#             with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
#                 sheet_name = "Results"
#                 df.to_excel(writer, index=False, sheet_name=sheet_name)
#                 workbook = writer.book
#                 sheet = workbook[sheet_name]

#                 # Styling headers
#                 header_font = Font(bold=True, color="FFFFFF")
#                 header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#                 border = Border(
#                     left=Side(border_style="thin", color="D3D3D3"),
#                     right=Side(border_style="thin", color="D3D3D3"),
#                     top=Side(border_style="thin", color="D3D3D3"),
#                     bottom=Side(border_style="thin", color="D3D3D3"),
#                 )

#                 # Apply header styles
#                 for col_num, col_name in enumerate(df.columns, 1):
#                     cell = sheet.cell(row=1, column=col_num)
#                     cell.font = header_font
#                     cell.fill = header_fill
#                     cell.border = border
#                     col_letter = get_column_letter(col_num)
#                     sheet.column_dimensions[col_letter].width = max(15, len(col_name) + 2)

#                 # Apply borders + center align
#                 for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
#                     for cell in row:
#                         cell.border = border
#                         cell.alignment = Alignment(vertical="center")

#             # Step 8: Create preview HTML (top 10)
#             preview_df = df.head(10)
#             html_table = preview_df.to_html(
#                 classes="dataframe table table-bordered table-hover",
#                 index=False,
#                 border=0,
#                 justify="center",
#                 na_rep="—"
#             )

#             preview_html = f"""
#             <div style="margin-top:10px; background:#f9f9f9; border-radius:8px; padding:10px;">
#                 <h4>📊 Preview (Top 10 of {total_rows} rows)</h4>
#                 {html_table}
#                 <div style="margin-top:10px; text-align:center;">
#                     <button 
#                         onclick="window.open('/downloads/{file_name}', '_blank')" 
#                         style="background-color:#0078D4; color:white; padding:8px 15px; border:none; border-radius:6px; cursor:pointer;">
#                         ⬇️ Download Excel (.xlsx)
#                     </button>
#                 </div>
#             </div>
#             """

#             yield {
#                 "event": "rows_preview",
#                 "data": {
#                     "columns": columns,
#                     "preview_html": preview_html,
#                     "row_count": total_rows,
#                     "file_name": file_name,
#                 },
#             }

#             # Step 9: Generate AI summary
#             yield {"event": "status", "message": "🤖 Summarizing query results..."}
#             rows_tuples = [tuple(r) for r in rows]
#             summary_text = await self.summarizer.summarize(user_query, fixed_sql, columns, rows_tuples, total_rows)
#             summary_text = summary_text or f"✅ Query executed successfully with {total_rows} rows."

#             await self._save_chat_message("assistant", summary_text)
#             await self._update_session_summary(summary_text)
#             await self._update_user_memory(summary_text)

#             exec_time_ms = int((time.time() - start_time) * 1000)
#             await self._log_query(user_query, fixed_sql, "valid", "success", None, total_rows, exec_time_ms)

#             yield {
#                 "event": "complete",
#                 "data": {
#                     "summary_html": f"<div><strong>AI Summary:</strong><p>{summary_text}</p></div>",
#                     "execution_time_ms": exec_time_ms,
#                     "row_count": total_rows,
#                     "sql": fixed_sql,
#                     "download_url": f"/downloads/{file_name}"
#                 },
#             }

#         except Exception as e:
#             logger.error(f"Stream process error: {e}")
#             await self._log_query(user_query, sql or "", "error", "failed", str(e))
#             yield {"event": "error", "message": str(e)}

#     # ============================================================
#     # 💾 Chat History / Logging Helpers
#     # ============================================================
#     async def _save_chat_message(self, role: str, content: str):
#         message_id = str(uuid.uuid4())
#         await self.session.execute(
#             text("""
#                 INSERT INTO public.chat_history (session_id, message_id, role, content, meta_data)
#                 VALUES (:sid, :mid, :r, :c, :meta)
#             """),
#             {
#                 "sid": self.session_id,
#                 "mid": message_id,
#                 "r": role,
#                 "c": content,
#                 "meta": json.dumps({"source": "ws"})
#             }
#         )
#         await self.session.commit()
#         logger.info(f"💾 Saved message {message_id} for session {self.session_id}")

#     async def _update_session_summary(self, new_summary: str):
#         await self.session.execute(text("""
#             INSERT INTO public.session_summaries (session_id, user_id, summary)
#             VALUES (:sid, NULL, :summary)
#             ON CONFLICT (session_id, user_id)
#             DO UPDATE SET summary = :summary, updated_at = CURRENT_TIMESTAMP
#         """), {"sid": self.session_id, "summary": new_summary})
#         await self.session.commit()

#     async def _update_user_memory(self, latest_summary: str):
#         if not self.user_id:
#             return
#         await self.session.execute(text("""
#             INSERT INTO public.user_memory (user_id, memory_summary)
#             VALUES (:uid, :mem)
#             ON CONFLICT (user_id)
#             DO UPDATE SET memory_summary = EXCLUDED.memory_summary, updated_at = CURRENT_TIMESTAMP
#         """), {"uid": self.user_id, "mem": latest_summary})
#         await self.session.commit()

#     async def _log_query(self, user_query, generated_sql=None, validation_status=None,
#                          execution_status=None, error_message=None, row_count=None, execution_time_ms=None):
#         await self.session.execute(text("""
#             INSERT INTO public.query_logs 
#             (session_id, user_query, generated_sql, validation_status, execution_status, error_message, row_count, execution_time_ms)
#             VALUES (:sid, :uq, :gsql, :v, :e, :err, :rows, :time)
#         """), {
#             "sid": self.session_id,
#             "uq": user_query,
#             "gsql": generated_sql,
#             "v": validation_status,
#             "e": execution_status,
#             "err": error_message,
#             "rows": row_count,
#             "time": execution_time_ms
#         })
#         await self.session.commit()



