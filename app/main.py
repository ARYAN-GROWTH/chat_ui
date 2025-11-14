
import os
import re
import json
import time
import uuid
import asyncio
import datetime
from typing import Any, Dict, List, Optional, Tuple
import logging
from logging.handlers import RotatingFileHandler

import asyncpg
from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast

# LangChain/OpenAI wrapper (your environment used this)
from langchain_openai import ChatOpenAI

# --- Load environment ---
load_dotenv()

# --- Logging setup ---
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)
logger = logging.getLogger("diamond-ai")
logger.setLevel(logging.INFO)
file_handler = RotatingFileHandler(f"{LOG_DIR}/server.log", maxBytes=5_000_000, backupCount=5)
file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
logger.addHandler(file_handler)
logger.addHandler(console_handler)
logger.info("Server Booting...")

# --- Environment helpers ---
def env(key: str, default: Optional[str] = None) -> str:
    val = os.getenv(key, default)
    if val is None:
        logger.error(f"Missing environment variable: {key}")
        raise RuntimeError(f"{key} missing")
    return val

OPENAI_API_KEY = env("OPENAI_API_KEY")
DATABASE_URL_ASYNC = env("DATABASE_URL_ASYNC")
SCHEMA = env("SCHEMA", "public")
TABLE_NAME = env("TABLE_NAME", "dev_diamond2")
SCHEMA_CACHE_TTL = int(env("SCHEMA_CACHE_TTL", "600"))
LAST_N_MEMORY = int(env("LAST_N_MEMORY", "3"))
SQL_MODEL_NAME = env("SQL_MODEL_NAME", "gpt-4.1-mini")
ANSWER_MODEL_NAME = env("ANSWER_MODEL_NAME", "gpt-4o-mini")
BASE_URL = os.getenv("NEXT_PUBLIC_BASE_URL", "").rstrip("/")

# Downloads folder (static)
DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# --- FastAPI app ---
app = FastAPI(title="Diamond DB AI Backend — Final")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"]
)

app.mount("/downloads", StaticFiles(directory=DOWNLOAD_DIR), name="downloads")

# --- Global exception handlers ---
@app.exception_handler(Exception)
async def global_handler(request, exc):
    logger.exception(f"Unhandled Error: {exc}")
    return JSONResponse({"error": "Internal Server Error"}, status_code=500)

@app.exception_handler(RequestValidationError)
async def validation_handler(request, exc):
    logger.error(f"Validation Error: {exc}")
    return JSONResponse({"error": "Invalid Input"}, status_code=422)

# --- DB pool and cache ---
_pg_pool: Optional[asyncpg.pool.Pool] = None
_schema_cache: Dict[str, Any] = {"value": None, "expires_at": 0}
session_memory: Dict[str, List[Dict[str, Any]]] = {}
session_lock = asyncio.Lock()

# --- LLM models (LangChain wrapper) ---
SQL_MODEL = ChatOpenAI(model=SQL_MODEL_NAME, temperature=0)
ANSWER_MODEL = ChatOpenAI(model=ANSWER_MODEL_NAME, temperature=0)

# --- Column normalization map (expand as needed) ---
# Map many user phrases/variants to canonical column names in DB
CANONICAL_COLUMNS = {
    "item_no": "item_no",
    "itemno": "item_no",
    "item": "item_no",
    "image": "image",
    "image_url": "image",
    "date": "date",
    "company": "company",
    "group": "group_name",
    "group_name": "group_name",
    "customer": "customer_name",
    "customername": "customer_name",
    "customer_name": "customer_name",
    "jgroup": "jgroup",
    "retail_range": "retail_range",
    "retail range": "retail_range",
    "range": "range_name",
    "main category": "main_category",
    "main_category": "main_category",
    "maincategory": "main_category",
    "subcategory": "subcategory",
    "sub cat": "subcategory",
    "subcat1": "subcategory",
    "collection": "collection",
    "division": "division",
    "diamond_ctw_fraction": "diamond_ctw_fraction",
    "diamond_ctw_range": "diamond_ctw_range",
    "diamond ctw range": "diamond_ctw_range",
    "custom_sd_ctrshap": "custom_sd_ctrshap",
    "custom_sd_ctrdesc": "custom_sd_ctrdesc",
    "sdc_mis_item_status": "sdc_mis_item_status",
    "new_tag": "new_tag",
    "secondary_sales_qty": "secondary_sales_qty",
    "secondary sales qty": "secondary_sales_qty",
    "secondary_sales_total_cost": "secondary_sales_total_cost",
    "secondary sales total cost": "secondary_sales_total_cost",
    "secondary_sales_value": "secondary_sales_value",
    "inventory_qty_final": "inventory_qty_final",
    "inventory qty": "inventory_qty_final",
    "inventory_qty": "inventory_qty_final",
    "inventory_cost_final": "inventory_cost_final",
    "inventory cost": "inventory_cost_final",
    "open_memo_qty": "open_memo_qty",
    "open_memo_amount": "open_memo_amount",
    "open_order_qty_asset": "open_order_qty_asset",
    "open_order_amount_asset": "open_order_amount_asset",
    "open_order_qty_memo": "open_order_qty_memo",
    "open_order_amount_memo": "open_order_amount_memo",
    
}



# --- Helpers ---
def safe_to_text(resp: Any) -> str:
    """
    Extract raw text from LLM response objects safely.
    """
    try:
        if resp is None:
            return ""
        if isinstance(resp, str):
            return resp
        content = getattr(resp, "content", None)
        if isinstance(content, str):
            return content
        if isinstance(resp, dict):
            if "content" in resp and isinstance(resp["content"], str):
                return resp["content"]
            if "messages" in resp and isinstance(resp["messages"], list) and resp["messages"]:
                last = resp["messages"][-1]
                if isinstance(last, dict) and "content" in last:
                    return last["content"]
            return ""
        if isinstance(resp, list) and resp:
            last = resp[-1]
            last_content = getattr(last, "content", None)
            if isinstance(last_content, str):
                return last_content
            if isinstance(last, dict) and "content" in last:
                return last["content"]
            return ""
        return str(resp)
    except Exception:
        return ""

def serialize_value(val: Any) -> Any:
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    if isinstance(val, datetime.date):
        return val.isoformat()
    return val

def serialize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    return {k: serialize_value(v) for k, v in row.items()}

# --- Postgres pool & schema retrieval ---
async def get_pg_pool() -> asyncpg.pool.Pool:
    global _pg_pool
    if _pg_pool is None:
        logger.info("Creating PostgreSQL Pool...")
        _pg_pool = await asyncpg.create_pool(DATABASE_URL_ASYNC, min_size=1, max_size=10)
    return _pg_pool

async def get_schema() -> List[Dict[str, Any]]:
    now = time.time()
    cached = _schema_cache.get("value")
    if cached and _schema_cache.get("expires_at", 0) > now:
        return cached

    pool = await get_pg_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema=$1 AND table_name=$2
            ORDER BY ordinal_position
            """,
            SCHEMA, TABLE_NAME
        )
    columns = [dict(r) for r in rows]
    _schema_cache.update({"value": columns, "expires_at": now + SCHEMA_CACHE_TTL})
    return columns

# --- SQL prompt template ---
SQL_PROMPT_TEMPLATE = r"""
You are a production-grade PostgreSQL query generator. Convert the user's natural language
request into ONE valid, safe, schema-restricted PostgreSQL SELECT statement.
Return only the SELECT statement (no explanations, no markdown, no backticks).

Placeholders:
- {schema}, {table}, {columns}, {memory}, {question}

Rules:
- Use ONLY columns listed in {columns} (exactly these names).
- Wrap column names in double quotes.
- If the question contains an item identifier (SKU-like), treat as item-specific lookup:
    - use WHERE "item_no" = '<ITEM>' and return item-specific rows.
    - Always include "item_no" in the SELECT if you filter by it.
- If user asked for a list (phrases: "list", "distinct", "all", "give me all", "show all"):
    - return SELECT DISTINCT "column" FROM ...
- Do NOT use other SQL statements (UPDATE/DELETE/INSERT/etc).
- Do NOT add comments or semicolons.
- Do NOT invent column names.
- Only add LIMIT if the user explicitly asked for one.

Examples:
User: "main category of 202SSBN129D_18LAW.C.J"
Output:
SELECT "item_no","main_category"
FROM {schema}.{table}
WHERE "item_no"='202SSBN129D_18LAW.C.J'

User: "inventory_cost_final and secondary_sales_total_cost of ITEM123"
Output:
SELECT "item_no","inventory_cost_final","secondary_sales_total_cost"
FROM {schema}.{table}
WHERE "item_no"='ITEM123'

Now generate the SELECT for:
Question: {question}
"""

# --- sanitize & clean helpers ---
def sanitize_sql(raw: str) -> str:
    """
    Extract a single SELECT statement and clean obvious escape artifacts.
    """
    if not raw:
        return "SELECT 1"
    txt = raw.strip().replace("```sql", "").replace("```", "")
    # extract up to first semicolon or end
    m = re.search(r"(select[\s\S]*?)(;|$)", txt, flags=re.I)
    if not m:
        raise ValueError("No SELECT found in SQL model output.")
    sql = m.group(1).strip()
    # fix common LLM escaping
    sql = sql.replace("\\'", "'")
    sql = sql.replace("\\\\", "\\")
    # reduce doubled single quotes to single
    sql = re.sub(r"'{2,}", "'", sql)
    return sql

def clean_item_literal(item: str) -> str:
    # remove slashes, unicode escapes, control chars, trailing quotes, and non-allowed characters
    if item is None:
        return ""
    s = item.replace("\\", "")
    s = re.sub(r"\\u[0-9A-Fa-f]{4}", "", s)
    s = "".join(ch for ch in s if ch.isprintable())
    s = s.rstrip("'").rstrip('"').rstrip("/")
    s = re.sub(r"[^A-Za-z0-9._-]", "", s)
    return s

def fix_itemno_quotes(sql: str) -> str:
    """
    Finds WHERE "item_no" = '...' pattern, cleans the literal, reconstructs.
    """
    try:
        pattern = r'WHERE\s+"item_no"\s*=\s*\'([^\']*)\''
        m = re.search(pattern, sql, flags=re.I)
        if not m:
            return sql
        raw_item = m.group(1)
        cleaned = clean_item_literal(raw_item)
        fixed = f'WHERE "item_no"=\'{cleaned}\''
        sql = re.sub(pattern, fixed, sql, flags=re.I)
        return sql
    except Exception:
        return sql

def enforce_item_no_in_select(sql: str) -> str:
    """
    Ensure queries that filter by item_no always select item_no in the result.
    """
    try:
        if re.search(r'WHERE\s+"item_no"', sql, flags=re.I):
            if re.search(r"SELECT\s+\*", sql, flags=re.I):
                return sql
            # Already has item_no?
            if re.search(r'\"item_no\"', sql):
                return sql
            # insert item_no right after SELECT
            return re.sub(r"SELECT\s+", 'SELECT "item_no", ', sql, flags=re.I, count=1)
        return sql
    except Exception:
        return sql

# --- Fallback intent detector (simple) ---
def detect_intent_fallback(question: str, schema_cols: List[Dict[str, Any]]) -> Optional[str]:
    q = question.lower()
    cols = [c["column_name"] for c in schema_cols]
    if "columns" in q or "database" in q:
        limited = ", ".join(f"\"{c}\"" for c in cols[:20])
        return f"SELECT {limited} FROM {SCHEMA}.{TABLE_NAME} LIMIT 1"
    # if question asks for main category broadly
    if "main category" in q and "main_category" in cols:
        return f"SELECT DISTINCT \"main_category\" FROM {SCHEMA}.{TABLE_NAME} LIMIT 200"
    if "subcategory" in q and "subcategory" in cols:
        return f"SELECT DISTINCT \"subcategory\" FROM {SCHEMA}.{TABLE_NAME} LIMIT 200"
    return None

# --- SQL generation using LLM ---
async def generate_sql(session_id: str, question: str) -> str:
    try:
        schema_cols = await get_schema()
    except Exception:
        return "SELECT 1"

    fallback = detect_intent_fallback(question, schema_cols)
    if fallback:
        logger.info(f"Using fallback SQL for question: {question}")
        return fallback

    async with session_lock:
        mem = session_memory.get(session_id, [])
    memory_text = "; ".join(f"Q:{p['q']} A:{p['a']}" for p in mem[-LAST_N_MEMORY:])

    columns_list = ", ".join(c["column_name"] for c in schema_cols)
    prompt = SQL_PROMPT_TEMPLATE.format(
        schema=SCHEMA,
        table=TABLE_NAME,
        columns=columns_list,
        memory=memory_text,
        question=question
    )

    try:
        # Call LLM off main loop
        resp = await asyncio.to_thread(SQL_MODEL.invoke, prompt)
        raw_sql = safe_to_text(resp)
        sql = sanitize_sql(raw_sql)
        sql = fix_itemno_quotes(sql)
        sql = enforce_item_no_in_select(sql)
        logger.info(f"Generated SQL: {sql}")
        return sql
    except Exception as e:
        logger.exception(f"SQL generation failed: {e}")
        return "SELECT 1"

# --- DB execution ---
async def exec_sql(sql: str) -> List[Dict[str, Any]]:
    pool = await get_pg_pool()
    try:
        async with pool.acquire() as conn:
            await conn.execute("SET statement_timeout='10000'")
            rows = await conn.fetch(sql)
            return [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"SQL Execution error: {e}\nSQL: {sql}")
        return []

# --- Answer generation (strict, uses only rows) ---
ANSWER_PROMPT = r"""
You are a strict data assistant. Given question: {question} and rows: {rows}
Return a short Markdown answer using ONLY the data present in rows. If rows is empty return:
**I’m sorry, but I couldn't find the information you requested.**
Do not mention SQL or databases.
"""

async def generate_answer(question: str, sql: str, rows: List[Dict[str, Any]]) -> str:
    rows_json = json.dumps(rows[:200], default=str)
    prompt = ANSWER_PROMPT.format(question=question, rows=rows_json)
    try:
        resp = await asyncio.to_thread(ANSWER_MODEL.invoke, prompt)
        return safe_to_text(resp)
    except Exception as e:
        logger.exception(f"Answer generation failed: {e}")
        # fallback: minimal strict output
        if not rows:
            return "I’m sorry, but I couldn't find the information you requested.\n\nYou may also ask:\n• What are the available categories?\n• Show details of a specific item."
        # produce a simple mapping output
        title = f"Results for your question: {question}"
        lines = [title]
        for r in rows[:10]:
            lines.append(json.dumps(serialize_row(r)))
        return "\n".join(lines)

# --- Excel writer (Pylance-safe) ---
def try_parse_date(value: Any) -> Any:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return value

def write_styled_excel_and_preview(rows: List[Dict[str, Any]], filename: str) -> Dict[str, Any]:
    file_path = os.path.join(DOWNLOAD_DIR, filename)
    wb = Workbook()
    sheet = cast(Worksheet, wb.active)
    sheet.title = "Results"

    if rows:
        # Determine headers from first row keys (stable order)
        headers = list(rows[0].keys())

        # Header row
        for idx, h in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=idx, value=h)
            # set safe style only if cell exists
            try:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E4F9C")
                cell.alignment = Alignment(horizontal="center")
            except Exception:
                pass

        # Data rows
        for r_i, row in enumerate(rows, start=2):
            for c_i, h in enumerate(headers, start=1):
                v = row.get(h)
                v_out = try_parse_date(v)
                sheet.cell(row=r_i, column=c_i, value=v_out)

        # Auto-width: iterate columns safely
        for col_i in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_i)
            max_len = 10
            try:
                for cell in sheet[col_letter]:
                    if cell is not None and cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_len + 4
            except Exception:
                # fallback ignore auto width issues
                pass

    wb.save(file_path)

    # Build URL
    if BASE_URL:
        # Ensure no trailing slash duplication
        file_url = f"{BASE_URL}/downloads/{filename}"
    else:
        file_url = f"/downloads/{filename}"

    return {
        "filename": filename,
        "preview_rows": [serialize_row(r) for r in rows[:8]],
        "file_url": file_url
    }

# --- Field extraction from question ---
def extract_requested_fields(question: str, available_cols: List[str]) -> List[str]:
    """
    Very simple heuristic: look for canonical names or synonyms in question text.
    Return list of canonical DB column names that we should select.
    Always prefer exact matches of canonical or synonym.
    """
    q = question.lower()
    found: List[str] = []

    # first check exact canonical mentions
    for syn, canonical in CANONICAL_COLUMNS.items():
        if syn in q and canonical in available_cols and canonical not in found:
            found.append(canonical)

    # fallback: check available_cols words
    for col in available_cols:
        if col.lower() in q and col not in found:
            found.append(col)

    return found

# --- Websocket endpoint ---
GREETINGS = {"hello", "hi", "hey", "namaste", "hola", "hlo"}

@app.websocket("/ws/query")
async def ws_query(ws: WebSocket):
    await ws.accept()
    session_id = uuid.uuid4().hex
    logger.info(f"Client Connected: {session_id}")
    async with session_lock:
        session_memory[session_id] = []

    try:
        while True:
            raw = await ws.receive_text()

            # ignore empty frames
            if not raw or raw.strip() in ["{}", "null", ""]:
                continue

            try:
                payload = json.loads(raw)
            except Exception:
                await ws.send_json({"type": "status", "message": "Invalid JSON"})
                continue

            question = str(payload.get("question", "")).strip()
            if not question:
                continue

            # Greeting
            if question.lower() in GREETINGS:
                await ws.send_json({"type": "answer", "message": "Hello! How can I help?"})
                continue

            await ws.send_json({"type": "status", "message": "Working..."})

            # Generate SQL
            sql = await generate_sql(session_id, question)

            # Execute SQL
            rows = await exec_sql(sql)

            # Make sure Excel includes item_no if the question is item-specific.
            # Also ensure Excel includes requested columns (even if not present in SQL result)
            schema_cols = await get_schema()
            available_cols = [c["column_name"] for c in schema_cols]

            # Determine requested fields heuristically
            requested = extract_requested_fields(question, available_cols)

            # If SQL filtered by item_no, ensure item_no present in SELECT/rows
            is_item_query = bool(re.search(r'WHERE\s+"item_no"\s*=', sql, flags=re.I))
            if is_item_query and "item_no" not in (rows[0].keys() if rows else []):
                # If rows are present but missing item_no, attempt to fetch full rows (SELECT * WHERE item_no=...)
                try:
                    item_m = re.search(r'WHERE\s+"item_no"\s*=\s*\'([^\']*)\'', sql, flags=re.I)
                    if item_m:
                        item_val = clean_item_literal(item_m.group(1))
                        full_sql = f'SELECT * FROM {SCHEMA}.{TABLE_NAME} WHERE "item_no" = \'{item_val}\''
                        rows = await exec_sql(full_sql)
                except Exception:
                    pass

            # Generate AI answer (strict)
            try:
                answer = await generate_answer(question, sql, rows)
            except Exception as e:
                logger.exception(f"Answer generation error: {e}")
                answer = "Could not generate answer."

            await ws.send_json({"type": "answer", "message": answer})

            
            if rows:
                # Build headers order
                first_cols: List[str] = []
                if is_item_query:
                    first_cols.append("item_no")
                # include requested unique canonical columns (preserve order & existence)
                for rc in requested:
                    if rc not in first_cols and rc in rows[0].keys():
                        first_cols.append(rc)
                # Ensure all columns from rows are present (append remaining)
                all_cols = list(rows[0].keys())
                ordered_cols = []
                for c in first_cols:
                    if c in all_cols:
                        ordered_cols.append(c)
                for c in all_cols:
                    if c not in ordered_cols:
                        ordered_cols.append(c)
                # Reorder each row to match ordered_cols
                ordered_rows: List[Dict[str, Any]] = []
                for r in rows:
                    new_r = {c: r.get(c) for c in ordered_cols}
                    ordered_rows.append(new_r)

                filename = f"query_{uuid.uuid4().hex[:8]}.xlsx"
                try:
                    res = write_styled_excel_and_preview(ordered_rows, filename)
                    await ws.send_json({
                        "type": "file",
                        "filename": filename,
                        "url": res["file_url"]
                    })
                    await ws.send_json({
                        "type": "file_preview",
                        "preview_rows": res["preview_rows"]
                    })
                except Exception as e:
                    logger.exception(f"Excel creation error: {e}")

            # Save memory
            async with session_lock:
                session_memory[session_id].append({"q": question, "a": answer})
                session_memory[session_id] = session_memory[session_id][-LAST_N_MEMORY:]

    except WebSocketDisconnect:
        logger.info(f"Client Disconnected: {session_id}")
        async with session_lock:
            session_memory.pop(session_id, None)
    except Exception as e:
        logger.exception(f"Unexpected websocket error: {e}")
        async with session_lock:
            session_memory.pop(session_id, None)
        try:
            await ws.close()
        except Exception:
            pass

# --- Root endpoint ---
@app.get("/")
async def root():
    return {"status": "ok", "msg": "Diamond AI Backend Running"}

# --- Startup / shutdown ---
@app.on_event("startup")
async def startup():
    logger.info("Server Started")

@app.on_event("shutdown")
async def shutdown():
    logger.info("Server Shutdown Started")
    global _pg_pool
    if _pg_pool:
        await _pg_pool.close()
        logger.info("PostgreSQL Pool Closed")
        _pg_pool = None
